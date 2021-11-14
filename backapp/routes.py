from sqlalchemy.orm import scoped_session, sessionmaker
from flask import render_template, url_for, jsonify, abort, request
from backapp.model import egrul, geoobject
from backapp import engine
from openpyxl import load_workbook
from backapp import app

def mserialize(val):
    #print(val)
    if type(val) == dict:
        return val
    else:
        return {v: getattr(val, v) for v in val.__dict__ if v!="_sa_instance_state"}

# mserialize - serialize val object
def mserialize_list(list):
    return [mserialize(m) for m in list]

@app.route('/map')
def create_map():
    return render_template("map.html")

# @app.route('/map')
# def create_map():
#     return render_template("map.html")

@app.route('/get_geo_object', methods=['GET', 'POST'])
def getGeoObject():
    data = request.get_json()
    Session = scoped_session(sessionmaker(autocommit=False, autoflush=False, bind=engine))
    #

    go = Session.query(geoobject).filter(geoobject.lat > data['min_lat'], geoobject.lat < data['max_lat'],
                                         geoobject.lon > data['min_lon'], geoobject.lon < data['max_lon']).all()
    Session.close()
    print(len(go))
    return jsonify(mserialize_list(go))


@app.route('/uploadegrul')
def uploadegrul():
    Session = scoped_session(sessionmaker(autocommit=False, autoflush=False, bind=engine))
    wb = load_workbook('C:\\Temp\\1.xlsx')
    source = wb.active
    target = wb.copy_worksheet(source)
    print(wb.get_sheet_names())
    data = []
    for x in range(2, target.max_row):
         data.append(egrul(title=target['A'+str(x)].value, inn=target['B'+str(x)].value, kpp=target['C'+str(x)].value,
                           adress=target['D'+str(x)].value, l_surname=target['E'+str(x)].value, l_name=target['F'+str(x)].value,
                           l_secondname=target['G'+str(x)].value,buis_type=target['H'+str(x)].value,phone=target['I'+str(x)].value,
                           email=target['J'+str(x)].value, profit=target['K'+str(x)].value,buis_price=target['L'+str(x)].value, edo=target['M'+str(x)].value,
                           reg_num_pf=target['O'+str(x)].value, site=target['Q'+str(x)].value))
    Session.add_all(data)
    Session.commit()
    Session.close()
    return '42'
