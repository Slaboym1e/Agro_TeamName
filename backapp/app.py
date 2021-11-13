from flask import Flask
from sqlalchemy.orm import scoped_session, sessionmaker
from flask import render_template, url_for, jsonify, abort
from backapp.model import egrul
from backapp.__inti__ import engine
from openpyxl import load_workbook
app = Flask(__name__)


@app.route('/')
def index():
    return 'AgroHacaton 2021!'

@app.route('/map')
def create_map():
    return render_template("map.html")

@app.route('/uploadegrul')
def uploadegrul():
    Session = scoped_session(sessionmaker(autocommit=False, autoflush=False, bind=engine))
    wb = load_workbook('C:\\Temp\\1.xlsx')
    source = wb.active
    target = wb.copy_worksheet(source)
    print(wb.get_sheet_names())
    data = []
    for x in range(2, target.max_row):
         data.append(egrul(title=target['A'+str(x)].value, inn=target['B'+str(x)].value, kpp=target['C'+str(x)].value,
                           adress=target['D'+str(x)].value, l_surname=target['E'+str(x)].value, l_name=target['F'+str(x)].value,
                           l_secondname=target['G'+str(x)].value,buis_type=target['H'+str(x)].value,phone=target['I'+str(x)].value,
                           email=target['J'+str(x)].value, profit=target['K'+str(x)].value,buis_price=target['L'+str(x)].value, edo=target['M'+str(x)].value,
                           reg_num_pf=target['O'+str(x)].value, site=target['Q'+str(x)].value))
    Session.add_all(data)
    Session.commit()
    Session.close()
    return '42'
